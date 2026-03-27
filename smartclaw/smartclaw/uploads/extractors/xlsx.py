"""XLSX extractor."""

from __future__ import annotations

from io import BytesIO

from smartclaw.uploads.models import ExtractionResult


class XlsxExtractor:
    """Extract readable worksheet text from XLSX files when openpyxl is available."""

    def extract(self, content: bytes, *, filename: str, media_type: str) -> ExtractionResult:
        try:
            from openpyxl import load_workbook
        except ImportError:
            return ExtractionResult(
                text="",
                summary=f"{filename} 是 XLSX 附件，但当前环境未安装 XLSX 解析依赖",
                supported=False,
            )

        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
            sheet_names: list[str] = []
            lines: list[str] = []
            row_count = 0
            for worksheet in workbook.worksheets:
                sheet_names.append(str(getattr(worksheet, "title", "Sheet")))
                lines.append(f"[Sheet] {sheet_names[-1]}")
                for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                    values = [str(value).strip() for value in row[:20] if value not in (None, "")]
                    if not values:
                        continue
                    lines.append("\t".join(values))
                    row_count += 1
                    if row_index >= 200:
                        lines.append("...")
                        break
            text = "\n".join(lines).strip()
        except Exception as exc:
            return ExtractionResult(
                text="",
                summary=f"{filename} XLSX 提取失败",
                supported=False,
                error=str(exc),
            )

        if not text:
            return ExtractionResult(
                text="",
                summary=f"{filename} 已上传，但未提取到表格内容",
                supported=False,
            )

        summary_parts = [f"{filename} 已上传"]
        if sheet_names:
            summary_parts.append("工作表: " + ", ".join(sheet_names[:4]))
        if row_count:
            summary_parts.append(f"提取行数约 {row_count}")
        return ExtractionResult(text=text, summary="；".join(summary_parts), supported=True)
